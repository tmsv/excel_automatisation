import openpyxl
wb = openpyxl.load_workbook('c:\\Users\\duvolle\\Downloads\\OpenDataScience\\excel_automatization\\FB Live Vids Onboarding (1).xlsx')

results = wb.create_sheet(title = 'Products')

j = 1

# for sheet in wb.worksheets:
# 	for i in range(1, sheet.max_row + 1):
# 		if sheet.cell(row = i, column = 3).value == 'no':
# 			results.cell(row = j, column = 1).value = sheet.cell(row = i, column = 2).value
# 			j += 1
# 		elif sheet.cell(row = i, column = 3).value == 'No':
# 			results.cell(row = j, column = 1).value = sheet.cell(row = i, column = 2).value
# 			j += 1

for sheet in wb.worksheets:
	# for i in range(1, sheet.max_row + 1):
	# 	results.cell(row = j, column = 1).value = sheet.cell(row = i, column = 12).value
	# 	j += 1
		# if sheet.cell(row = i, column = 12).value == True:
		# 	results.cell(row = j, column = 1).value = sheet.cell(row = i, column = 12).value
		# 	j += 1
		# elif sheet.cell(row = i, column = 3).value == False:
		# 	continue
			# results.cell(row = j, column = 1).value = sheet.cell(row = i, column = 2).value
			# j += 1
	for i in range(1, sheet.max_row + 1):
		if sheet.cell(row = i, column = 12).value is None:
			continue
		else:
			results.cell(row = j, column = 1).value = sheet.cell(row = i, column = 12).value
			j += 1


	#for sheet in wb.worksheets:
		#wb.remove_sheet(sheet)

wb.save('PRODUCTS - 09.09.2019.xlsx')
